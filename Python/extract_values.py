import openpyxl
import json
import os

def extract_values(directory):
    files = [f for f in os.listdir(directory) if f.endswith('.xlsx')]
    all_data = {}

    for file in files:
        path = os.path.join(directory, file)
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            file_data = {}
            target_sheets = [
                "0.结构参数", 
                "0.地勘及桩计算", 
                "1.U型槽主筋验算", 
                "2.U型槽钢筋", 
                "3.抗浮桩", 
                "3.抗浮桩钢筋计算", 
                "4.防水"
            ]
            for sheet_name in target_sheets:
                if sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    # Extract a block of data, e.g., 50 rows, 30 columns
                    data = []
                    for r in range(1, 51):
                        row = []
                        for c in range(1, 31):
                            val = sheet.cell(row=r, column=c).value
                            row.append(str(val) if val is not None else "")
                        data.append(row)
                    file_data[sheet_name] = data
            all_data[file] = file_data
        except Exception as e:
            print(f"Error processing {file}: {e}")
    
    with open("u_channel_data.json", "w", encoding="utf-8") as f:
        json.dump(all_data, f, ensure_ascii=False, indent=2)

if __name__ == "__main__":
    extract_values(r"h:\U形槽配筋软件\DesignSheet")
