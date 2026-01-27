import openpyxl
import os

def find_table(folder):
    files = [f for f in os.listdir(folder) if f.endswith('.xlsx')]
    for filename in files:
        path = os.path.join(folder, filename)
        try:
            wb = openpyxl.load_workbook(path, read_only=True)
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                # Search first few rows for the title
                for row in sheet.iter_rows(max_row=50, max_col=10):
                    for cell in row:
                        if cell.value == 2285 or cell.value == 2285.0:
                            print(f"Found 2285 in {filename} -> {sheet_name} at {cell.coordinate}")
                            return path, sheet_name
        except Exception as e:
            continue
    return None, None

folder = r"h:\U形槽配筋软件\DesignSheet"
path, sheet = find_table(folder)
if path:
    print(f"Target file: {path}")
    print(f"Target sheet: {sheet}")
else:
    print("Not found.")
