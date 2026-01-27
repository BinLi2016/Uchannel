import openpyxl
import os
import json

def extract_comprehensive_logic(file_path):
    print(f"Deep analyzing {file_path}...")
    try:
        # Load without data_only to see formulas
        wb = openpyxl.load_workbook(file_path, data_only=False, read_only=True)
        logic = {
            "file": os.path.basename(file_path),
            "sheets": {}
        }
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            formulas = {}
            # Increase limits to capture full sheet logic
            for row in sheet.iter_rows(max_row=300, max_col=50):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formulas[cell.coordinate] = cell.value
            
            logic["sheets"][sheet_name] = {
                "formula_count": len(formulas),
                "formulas": formulas
            }
            
        return logic
    except Exception as e:
        return {"error": str(e)}

def main():
    folder = r"h:\U形槽配筋软件\DesignSheet"
    files = [f for f in os.listdir(folder) if f.endswith('.xlsx')]
    
    all_results = {}
    for filename in files:
        path = os.path.join(folder, filename)
        all_results[filename] = extract_comprehensive_logic(path)
        
    with open("deep_excel_meta.json", "w", encoding="utf-8") as f:
        json.dump(all_results, f, ensure_ascii=False, indent=2)
    print("\n[COMPLETE] Deep metadata saved to deep_excel_meta.json")

if __name__ == "__main__":
    main()
