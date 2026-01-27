import openpyxl
import os
import json

def extract_table_data(file_path, sheet_name):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb[sheet_name]
    
    table_data = []
    # Find YU1 first
    start_row = 1
    start_col = 1
    found = False
    for row in sheet.iter_rows(max_row=100, max_col=20):
        for cell in row:
            if cell.value == "YU1":
                start_row = cell.row
                start_col = cell.column
                found = True
                break
        if found: break
    
    if not found:
        print("YU1 not found in detail.")
        return []

    # Read rows until we don't see YU...
    for r in range(start_row, start_row + 30):
        row_id = sheet.cell(row=r, column=start_col).value
        if not row_id or not str(row_id).startswith("YU"):
            if r > start_row: break
            continue
        
        row_data = []
        for c in range(start_col, start_col + 25):
            val = sheet.cell(row=r, column=c).value
            row_data.append(val)
        table_data.append(row_data)
    
    return table_data

path = r"h:\U形槽配筋软件\DesignSheet\U型槽施工图设计 振华路.xlsx"
sheet = "基坑计算"
data = extract_table_data(path, sheet)
print(json.dumps(data, indent=2, ensure_ascii=False))
