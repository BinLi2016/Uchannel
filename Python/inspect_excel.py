import openpyxl
import os
import json

def extract_logic(file_path):
    print(f"Processing {file_path}...")
    try:
        wb = openpyxl.load_workbook(file_path, data_only=False, read_only=True)
        logic = {
            "sheets": wb.sheetnames,
            "formulas": {}
        }
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            formulas = {}
            # Limit search to first 100 rows and 20 columns to find logic patterns
            for row in sheet.iter_rows(max_row=100, max_col=20):
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formula = cell.value
                        if formula not in formulas:
                            formulas[formula] = cell.coordinate
            
            logic["formulas"][sheet_name] = formulas
            
        return logic
    except Exception as e:
        return {"error": str(e)}

def main():
    folder = r"h:\U形槽配筋软件\DesignSheet"
    files = [f for f in os.listdir(folder) if f.endswith('.xlsx')]
    
    # Analyze a few representative files
    results = {}
    for i, filename in enumerate(files):
        if i >= 3: break # Limit to first 3 files for initial inspection
        path = os.path.join(folder, filename)
        results[filename] = extract_logic(path)
        
    with open("excel_meta.json", "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    print("Metadata saved to excel_meta.json")

if __name__ == "__main__":
    main()
